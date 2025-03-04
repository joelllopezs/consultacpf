import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# - Entrar na planilha e extrair o cpf
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/') 

# Abrir o arquivo de fechamento antes do loop
planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
pagina_fechamento = planilha_fechamento['Sheet1']

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True): 
    nome, valor, cpf, vencimento = linha

    # Entrar no site e usar o cpf daquele cliente para pesquisar o status de pagamento daquele cliente
    sleep(4)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    # Verificar se está em dia ou atrasado
    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(3)

    try:
        status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
        if status.text == 'em dia':
            # Se estiver em dia, pegar a data do pagamento  e o método de pagamento
            data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
            metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
            data_pagamento_limpo = data_pagamento.text.split()[3]
            metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

            pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
        else: 
            # Caso contrário , colocar status como pendente
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

        planilha_fechamento.save('planilha fechamento.xlsx')
    except Exception as e:
        print(f"Erro ao processar o CPF {cpf}: {e}")

driver.quit()  # Fechar o navegador após o processamento de todos os clientes
